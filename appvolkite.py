import streamlit as st
import pandas as pd
import time
import io
import requests
from kiteconnect import KiteConnect
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ──────────────────────────────────────────────
# STREAMLIT PAGE INITIALIZATION & THEME CSS
# ──────────────────────────────────────────────
st.set_page_config(layout="wide", page_title="Pro Option Terminal", page_icon="📊")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght=400;700;800&family=JetBrains+Mono:wght=500;700&display=swap');

/* Light Blue/Red Theme Background */
[data-testid="stAppViewContainer"], [data-testid="stHeader"], .main {
    background-color: #f0f6ff !important;
}

/* Dark Text for Light Background */
html, body, [class*="css"] {
    color: #0d1b2a !important;
    font-family: 'Inter', sans-serif !important;
}

/* Table Headers */
.section-headers { display: grid; grid-template-columns: 1fr 110px 1fr; gap: 10px; margin-bottom: 5px; }
.sh { 
    text-align: center; padding: 10px; font-weight: 700; border-radius: 4px; 
    border: 1px solid #b8d4f0; background: #daeaf8; font-size: 0.8rem; color: #0d1b2a;
}

/* Dataframe Container */
[data-testid="stDataFrameResizable"] {
    background-color: #f0f6ff !important;
    border: 1px solid #b8d4f0 !important;
}
.stDataFrame th {
    background-color: #c8dff5 !important;
    color: #1a3a5c !important;
    font-size: 0.7rem !important;
}

/* Style for Buttons */
div.stButton > button {
    width: 100%;
    background-color: #daeaf8;
    color: #0d1b2a;
    border: 1px solid #7ab3e0;
}
div.stButton > button:hover {
    border-color: #1976d2;
    color: #1976d2;
}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# SIDEBAR CREDENTIALS MANAGEMENT PANEL (CALLBACK FIX)
# ──────────────────────────────────────────────
st.sidebar.title("Credentials")

# Initialize structural variables in state safely
if "active_api_key" not in st.session_state:
    st.session_state["active_api_key"] = ""
if "active_access_token" not in st.session_state:
    st.session_state["active_access_token"] = ""

# 1. Bind inputs directly to session state
zerodha_api_key = st.sidebar.text_input(
    "Zerodha API key", 
    type="password", 
    key="active_api_key"
)

zerodha_access_token = st.sidebar.text_input(
    "Zerodha access token", 
    type="password", 
    key="active_access_token"
)

# ── THE FIX: CALLBACK FUNCTION ──
# This function runs BEFORE widgets are drawn, making state modifications legal
def token_generation_callback():
    secret = st.session_state.get("temp_secret", "")
    redirect_input = st.session_state.get("temp_redirect", "")
    api_key = st.session_state.get("active_api_key", "")
    
    if api_key and secret and redirect_input:
        try:
            req_token = redirect_input
            if "request_token=" in redirect_input:
                req_token = redirect_input.split("request_token=")[1].split("&")[0]
            
            # Generate session
            temp_kite = KiteConnect(api_key=api_key)
            session_data = temp_kite.generate_session(req_token, api_secret=secret)
            
            # Writing directly to the widget key is perfectly valid inside a callback!
            st.session_state["active_access_token"] = session_data["access_token"]
            st.session_state["token_success_msg"] = "✅ Access Token Generated successfully!"
            
            # Clear fields
            st.session_state["temp_redirect"] = ""
        except Exception as e:
            st.session_state["token_error_msg"] = f"❌ Generation Failed: {e}"
    else:
        st.session_state["token_error_msg"] = "❌ Missing Key, Secret, or Redirect URL."


# Expandable Token Generator Section
with st.sidebar.expander("🔄 Generate Zerodha access token", expanded=False):
    if st.button("Open Kite login"):
        if zerodha_api_key:
            login_url = f"https://kite.trade/connect/login?v=3&api_key={zerodha_api_key}"
            st.markdown(f"[👉 Click here to login to Kite]({login_url})")
        else:
            st.sidebar.error("Please enter your Zerodha API Key first!")

    # Separate temporary keys for secret and redirect URL to avoid conflicting with main token widget
    st.sidebar.text_input("Zerodha API secret", type="password", key="temp_secret")
    st.sidebar.text_input("Request token or full redirect URL", key="temp_redirect")

    # Link button to the callback routine
    st.button("Generate access token", on_click=token_generation_callback)

    # Show success/error banners from the callback execution
    if "token_success_msg" in st.session_state and st.session_state["token_success_msg"]:
        st.success(st.session_state["token_success_msg"])
        del st.session_state["token_success_msg"] # flash message
    if "token_error_msg" in st.session_state and st.session_state["token_error_msg"]:
        st.error(st.session_state["token_error_msg"])
        del st.session_state["token_error_msg"] # flash message

# Connection Verification Testing
if st.sidebar.button("Test Zerodha connection"):
    if zerodha_api_key and zerodha_access_token:
        try:
            test_kite = KiteConnect(api_key=zerodha_api_key)
            test_kite.set_access_token(zerodha_access_token)
            profile = test_kite.profile()
            st.sidebar.success(f"Connected! User: {profile.get('user_name', 'Active Session')}")
        except Exception as e:
            st.sidebar.error(f"Connection failed: {e}")
    else:
        st.sidebar.error("Provide both API Key and Access Token to test.")

# Static downstream configuration pipelines
TELEGRAM_TOKEN = st.sidebar.text_input("Telegram bot token", type="password", value="7851529826:AAHfyHVrVZi5iQubljaNgde76gPhr8pxql4")
TELEGRAM_CHAT_ID = st.sidebar.text_input("Telegram chat id", value="567677761")

# Global Active API Client mapping setup
kite = None
if zerodha_api_key and zerodha_access_token:
    try:
        kite = KiteConnect(api_key=zerodha_api_key)
        kite.set_access_token(zerodha_access_token)
    except Exception as e:
        st.error(f"Kite Initialization Matrix Fault: {e}")

# Index step definition maps
UNDERLYING_MAP = {
    "NIFTY":  {"UnderlyingSymbol": "NSE:NIFTY 50",  "ExchangeSymbol": "NIFTY",  "step": 50},
    "SENSEX": {"UnderlyingSymbol": "BSE:SENSEX",    "ExchangeSymbol": "SENSEX", "step": 100},
}

# ──────────────────────────────────────────────
# VISUALIZATION & TELEGRAM ALERT ENGINE
# ──────────────────────────────────────────────

def send_telegram_combined_analysis(index_name, ltp, atm, pcr, df, step):
    try:
        df = df[(df["STRIKE"] >= atm - step*5) & (df["STRIKE"] <= atm + step*5)]
        df = df.sort_values("STRIKE", ascending=False)

        width, height = 850, 80 + len(df)*60 + 40
        img = Image.new("RGB", (width, height), (10, 12, 18))
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()

        draw.text((20, 15), f"🚀 {index_name} DUAL ANALYSIS | LTP: {ltp:,.0f} | PCR: {pcr:.2f}", fill=(255,255,255), font=font)
        draw.text((20, 35), "LEFT: CALL (Vol/ΔOI) | RIGHT: PUT (Vol/ΔOI)", fill=(150, 150, 150), font=font)

        max_vol = max(df["_cv"].max(), df["_pv"].max(), 1)
        max_oi  = max(df["_cd"].abs().max(), df["_pd"].abs().max(), 1)

        y = 75
        bar_max_w = 220

        for _, r in df.iterrows():
            strike = int(r["STRIKE"])
            cv, pv = r["_cv"], r["_pv"]
            cd, pd = r["_cd"], r["_pd"]

            cv_w = int((cv / max_vol) * bar_max_w)
            pv_w = int((pv / max_vol) * bar_max_w)
            draw.rectangle([20, y, 20 + cv_w, y + 8], fill=(100, 116, 139))
            draw.rectangle([width - 20 - pv_w, y, width - 20, y + 8], fill=(100, 116, 139))

            cd_w = int((abs(cd) / max_oi) * bar_max_w)
            pd_w = int((abs(pd) / max_oi) * bar_max_w)
            
            ce_color = (239, 68, 68) if cd > 0 else (34, 197, 94)
            draw.rectangle([20, y + 12, 20 + cd_w, y + 25], fill=ce_color)
            
            pe_color = (34, 197, 94) if pd > 0 else (239, 68, 68)
            draw.rectangle([width - 20 - pd_w, y + 12, width - 20, y + 25], fill=pe_color)

            txt = f"{strike} ATM" if strike == atm else str(strike)
            if strike == atm:
                draw.rectangle([width//2 - 50, y, width//2 + 50, y + 25], outline=(255,255,255))
            
            draw.text((width//2 - 30, y + 5), txt, fill=(255, 255, 255), font=font)

            draw.text((20 + cv_w + 5, y), f"V:{cv/1e5:.1f}L", fill=(148, 163, 184), font=font)
            draw.text((width - 20 - pv_w - 70, y), f"V:{pv/1e5:.1f}L", fill=(148, 163, 184), font=font)
            draw.text((20 + cd_w + 5, y + 12), f"Δ:{cd/1e5:.1f}L", fill=ce_color, font=font)
            draw.text((width - 20 - pd_w - 70, y + 12), f"Δ:{pd/1e5:.1f}L", fill=pe_color, font=font)

            y += 55

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto",
                      data={"chat_id": TELEGRAM_CHAT_ID},
                      files={"photo": ("analysis.png", buf, "image/png")}, timeout=15)
    except Exception as e:
        print(f"Error generating dual chart: {e}")


def render_strikewise_image_streamlit(index_name, ltp, atm, pcr, df, step):
    try:
        df = df[(df["STRIKE"] >= atm - step*5) & (df["STRIKE"] <= atm + step*5)]
        df = df.sort_values("STRIKE", ascending=False)

        width, height = 800, 50 + len(df)*40 + 40
        img = Image.new("RGB", (width, height), (15, 18, 25))
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()

        draw.text((20, 10), f"{index_name} | LTP: {ltp:,.0f}", fill=(255,255,255), font=font)
        max_vol = max(df["_cv"].max(), df["_pv"].max(), 1)

        all_volumes = pd.concat([df["_cv"], df["_pv"]]).dropna().unique()
        all_volumes = sorted(all_volumes, reverse=True)
        
        top1 = all_volumes[0] if len(all_volumes) > 0 else -1
        top2 = all_volumes[1] if len(all_volumes) > 1 else -1
        top3 = all_volumes[2] if len(all_volumes) > 2 else -1

        BASE_GREEN, BASE_RED, BASE_GRAY = (34, 197, 94), (239, 68, 68), (156, 163, 175)
        TOP1_PINK, TOP2_ORANGE, TOP3_WHITE = (236, 72, 153), (249, 115, 22), (255, 255, 255)

        def determine_color(current_vol, opposing_vol):
            if current_vol == top1: return TOP1_PINK
            elif current_vol == top2: return TOP2_ORANGE
            elif current_vol == top3: return TOP3_WHITE
            if current_vol > opposing_vol: return BASE_GREEN
            elif opposing_vol > current_vol: return BASE_RED
            return BASE_GRAY

        y = 50
        bar_max_width = 200

        for _, r in df.iterrows():
            strike = int(r["STRIKE"])
            c_vol, p_vol = r["_cv"], r["_pv"]
            c_delta, p_delta = r["_cd"], r["_pd"]

            if c_delta > p_delta: color = (255, 80, 80)
            elif p_delta > c_delta: color = (80, 255, 120)
            else: color = (200, 200, 200)

            ce_color = determine_color(c_vol, p_vol)
            pe_color = determine_color(p_vol, c_vol)

            c_width = int((c_vol / max_vol) * bar_max_width)
            draw.rectangle([20, y, 20 + c_width, y + 15], fill=ce_color)

            p_width = int((p_vol / max_vol) * bar_max_width)
            draw.rectangle([width - 20 - p_width, y, width - 20, y + 15], fill=pe_color)

            txt = f"{strike} ATM" if strike == atm else f"{strike}"
            draw.text((width//2 - 40, y), txt, fill=color, font=font)

            draw.text((20 + c_width + 5, y), f"{c_vol/1e5:.1f}L", fill=(255,255,255), font=font)
            draw.text((width - 20 - p_width - 60, y), f"{p_vol/1e5:.1f}L", fill=(255,255,255), font=font)
            y += 35

        draw.text((20, height - 30), f"PCR: {pcr:.2f}", fill=(255,255,255), font=font)
        return img
    except Exception as e:
        st.error(f"Error rendering chart: {e}")
        return None


def send_telegram_strikewise_image(index_name, ltp, atm, pcr, df, step):
    try:
        df = df[(df["STRIKE"] >= atm - step*5) & (df["STRIKE"] <= atm + step*5)]
        df = df.sort_values("STRIKE", ascending=False)

        width, height = 800, 50 + len(df)*40 + 40
        img = Image.new("RGB", (width, height), (15, 18, 25))
        draw = ImageDraw.Draw(img)
        font = ImageFont.load_default()

        draw.text((20, 10), f"{index_name} | LTP: {ltp:,.0f}", fill=(255,255,255), font=font)
        max_vol = max(df["_cv"].max(), df["_pv"].max(), 1)

        all_volumes = pd.concat([df["_cv"], df["_pv"]]).dropna().unique()
        all_volumes = sorted(all_volumes, reverse=True)
        
        top1 = all_volumes[0] if len(all_volumes) > 0 else -1
        top2 = all_volumes[1] if len(all_volumes) > 1 else -1
        top3 = all_volumes[2] if len(all_volumes) > 2 else -1

        BASE_GREEN, BASE_RED, BASE_GRAY = (34, 197, 94), (239, 68, 68), (156, 163, 175)
        TOP1_PINK, TOP2_ORANGE, TOP3_WHITE = (236, 72, 153), (249, 115, 22), (255, 255, 255)

        def determine_color(current_vol, opposing_vol):
            if current_vol == top1: return TOP1_PINK
            elif current_vol == top2: return TOP2_ORANGE
            elif current_vol == top3: return TOP3_WHITE
            if current_vol > opposing_vol: return BASE_GREEN
            elif opposing_vol > current_vol: return BASE_RED
            return BASE_GRAY

        y = 50
        bar_max_width = 200

        for _, r in df.iterrows():
            strike = int(r["STRIKE"])
            c_vol, p_vol = r["_cv"], r["_pv"]
            c_delta, p_delta = r["_cd"], r["_pd"]

            if c_delta > p_delta: color = (255, 80, 80)
            elif p_delta > c_delta: color = (80, 255, 120)
            else: color = (200, 200, 200)

            ce_color = determine_color(c_vol, p_vol)
            pe_color = determine_color(p_vol, c_vol)

            c_width = int((c_vol / max_vol) * bar_max_width)
            draw.rectangle([20, y, 20 + c_width, y + 15], fill=ce_color)

            p_width = int((p_vol / max_vol) * bar_max_width)
            draw.rectangle([width - 20 - p_width, y, width - 20, y + 15], fill=pe_color)

            txt = f"{strike} ATM" if strike == atm else f"{strike}"
            draw.text((width//2 - 40, y), txt, fill=color, font=font)

            draw.text((20 + c_width + 5, y), f"{c_vol/1e5:.1f}L", fill=(255,255,255), font=font)
            draw.text((width - 20 - p_width - 60, y), f"{p_vol/1e5:.1f}L", fill=(255,255,255), font=font)
            y += 35

        draw.text((20, height - 30), f"PCR: {pcr:.2f}", fill=(255,255,255), font=font)

        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)

        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendPhoto",
            data={"chat_id": TELEGRAM_CHAT_ID},
            files={"photo": ("oc.png", buf, "image/png")},
            timeout=10
        )
    except Exception as e:
        pass


def send_telegram_alert(index_name, ltp, atm, expiry, pcr, df):
    try:
        max_c_oi_chg_row = df.loc[df["_cd"].idxmax()]
        max_c_vol_row    = df.loc[df["_cv"].idxmax()]
        max_p_oi_chg_row = df.loc[df["_pd"].idxmax()]
        max_p_vol_row    = df.loc[df["_pv"].idxmax()]

        c_arrow = "▲" if max_c_oi_chg_row["_cd"] >= 0 else "▼"
        p_arrow = "▲" if max_p_oi_chg_row["_pd"] >= 0 else "▼"

        msg = (
            f"📊 *{index_name} Option Chain Alert*\n"
            f"🕐 {time.strftime('%d-%b %H:%M')} | Expiry: {expiry}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"💰 LTP: `{ltp:,.0f}` | ATM: `{int(atm)}` | PCR: `{pcr:.2f}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📈 *CALL (CE)*\n"
            f"  🔹 Highest OI Chg : `{int(max_c_oi_chg_row['STRIKE'])}` — ΔOI: `{max_c_oi_chg_row['_cd']/1e5:.2f}L {c_arrow}` | LTP: `{max_c_oi_chg_row['C LTP']}`\n"
            f"  🔹 Highest Vol    : `{int(max_c_vol_row['STRIKE'])}` — Vol: `{max_c_vol_row['_cv']/1e5:.2f}L` | LTP: `{max_c_vol_row['C LTP']}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📉 *PUT (PE)*\n"
            f"  🔸 Highest OI Chg : `{int(max_p_oi_chg_row['STRIKE'])}` — ΔOI: `{max_p_oi_chg_row['_pd']/1e5:.2f}L {p_arrow}` | LTP: `{max_p_oi_chg_row['P LTP']}`\n"
            f"  🔸 Highest Vol    : `{int(max_p_vol_row['STRIKE'])}` — Vol: `{max_p_vol_row['_pv']/1e5:.2f}L` | LTP: `{max_p_vol_row['P LTP']}`\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"_Auto-alert on every page refresh_"
        )
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": msg, "parse_mode": "Markdown"},
            timeout=5
        )
    except Exception as e:
        pass


def send_excel_to_telegram(index_name, ltp, atm, expiry, pcr, df,
                            c_vol_top3, c_oi_top3, p_vol_top3, p_oi_top3,
                            min_c_oi_idx, min_p_oi_idx,
                            c_neg_oi_top3=None, p_neg_oi_top3=None):
    if c_neg_oi_top3 is None: c_neg_oi_top3 = []
    if p_neg_oi_top3 is None: p_neg_oi_top3 = []
    try:
        display_cols = ["C OI CH%","C VOL (L)","CALL OI (L)","C Δ OI","C LTP",
                        "STRIKE","IV","P LTP","P Δ OI","PUT OI (L)","P VOL (L)","P OI CH%"]
        export_df = df[display_cols].copy()

        wb = Workbook()
        ws = wb.active
        ws.title = f"{index_name} OC"

        def fill(hex_col): return PatternFill("solid", fgColor=hex_col.replace("#",""))

        FILLS = {
            "CYAN1":  fill("#1976d2"), "CYAN2":  fill("#64b5f6"), "CYAN3":  fill("#bbdefb"),
            "PINK1":  fill("#c62828"), "PINK2":  fill("#ef5350"), "PINK3":  fill("#ffcdd2"),
            "YELLOW": fill("#ffe082"), "YELLOW2": fill("#ffd54f"), "YELLOW3": fill("#fff9c4"), "WHITE":  fill("#ffffff"),
            "STRIKE": fill("#c8dff5"), "DARK":   fill("#f0f6ff"), "HEADER": fill("#daeaf8"),
        }
        WHITE_FONT  = Font(color="0D1B2A", bold=True, name="Calibri", size=10)
        BLACK_FONT  = Font(color="0D1B2A", bold=True, name="Calibri", size=10)
        NORMAL_FONT = Font(color="0D1B2A", name="Calibri", size=10)
        GREY_FONT   = Font(color="2C5F8A", name="Calibri", size=10)
        CENTER      = Alignment(horizontal="center", vertical="center")
        thin        = Side(style="thin", color="B8D4F0")
        BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells(f"A1:{get_column_letter(len(display_cols))}1")
        title_cell = ws["A1"]
        title_cell.value = (f"{index_name}  |  LTP: {ltp:,.0f}  |  ATM: {int(atm)}  "
                            f"|  PCR: {pcr:.2f}  |  Expiry: {expiry}  "
                            f"|  {time.strftime('%d-%b-%Y %H:%M')}")
        title_cell.fill, title_cell.font, title_cell.alignment = FILLS["HEADER"], Font(color="0D1B2A", bold=True, name="Calibri", size=11), CENTER
        ws.row_dimensions[1].height = 22

        for ci, col in enumerate(display_cols, 1):
            cell = ws.cell(row=2, column=ci, value=col)
            cell.fill, cell.font, cell.alignment, cell.border = FILLS["HEADER"], Font(color="1A3A5C", bold=True, name="Calibri", size=10), CENTER, BORDER
        ws.row_dimensions[2].height = 18

        for ri, (_, row) in enumerate(export_df.iterrows(), 3):
            ws.row_dimensions[ri].height = 18
            strike_val = df.loc[ri-3, "STRIKE"]
            is_atm     = (strike_val == atm)
            df_idx     = ri - 3

            for ci, col in enumerate(display_cols, 1):
                cell = ws.cell(row=ri, column=ci, value=row[col])
                cell.alignment, cell.border = CENTER, BORDER
                cell.fill = FILLS["WHITE"] if is_atm else FILLS["DARK"]
                cell.font = BLACK_FONT if is_atm else NORMAL_FONT

                if col == "STRIKE":
                    cell.fill = FILLS["WHITE"] if is_atm else FILLS["STRIKE"]
                    cell.font = BLACK_FONT if is_atm else WHITE_FONT
                elif col == "IV":
                    cell.fill, cell.font = FILLS["STRIKE"], GREY_FONT
                elif col == "C VOL (L)":
                    if df_idx == c_vol_top3[0]: cell.fill, cell.font = FILLS["CYAN1"], BLACK_FONT
                    elif len(c_vol_top3) > 1 and df_idx == c_vol_top3[1]: cell.fill, cell.font = FILLS["CYAN2"], WHITE_FONT
                    elif len(c_vol_top3) > 2 and df_idx == c_vol_top3[2]: cell.fill, cell.font = FILLS["CYAN3"], WHITE_FONT
                elif col == "C Δ OI":
                    if len(c_neg_oi_top3) > 2 and df_idx == c_neg_oi_top3[2]: cell.fill, cell.font = FILLS["YELLOW3"], BLACK_FONT
                    if len(c_neg_oi_top3) > 1 and df_idx == c_neg_oi_top3[1]: cell.fill, cell.font = FILLS["YELLOW2"], BLACK_FONT
                    if len(c_neg_oi_top3) > 0 and df_idx == c_neg_oi_top3[0]: cell.fill, cell.font = FILLS["YELLOW"], BLACK_FONT
                    elif df_idx == c_oi_top3[0]: cell.fill, cell.font = FILLS["CYAN1"], BLACK_FONT
                    elif len(c_oi_top3) > 1 and df_idx == c_oi_top3[1]: cell.fill, cell.font = FILLS["CYAN2"], WHITE_FONT
                    elif len(c_oi_top3) > 2 and df_idx == c_oi_top3[2]: cell.fill, cell.font = FILLS["CYAN3"], WHITE_FONT
                elif col == "P Δ OI":
                    if len(p_neg_oi_top3) > 2 and df_idx == p_neg_oi_top3[2]: cell.fill, cell.font = FILLS["YELLOW3"], BLACK_FONT
                    if len(p_neg_oi_top3) > 1 and df_idx == p_neg_oi_top3[1]: cell.fill, cell.font = FILLS["YELLOW2"], BLACK_FONT
                    if len(p_neg_oi_top3) > 0 and df_idx == p_neg_oi_top3[0]: cell.fill, cell.font = FILLS["YELLOW"], BLACK_FONT
                    elif df_idx == p_oi_top3[0]: cell.fill, cell.font = FILLS["PINK1"], BLACK_FONT
                    elif len(p_oi_top3) > 1 and df_idx == p_oi_top3[1]: cell.fill, cell.font = FILLS["PINK2"], WHITE_FONT
                    elif len(p_oi_top3) > 2 and df_idx == p_oi_top3[2]: cell.fill, cell.font = FILLS["PINK3"], WHITE_FONT
                elif col == "P VOL (L)":
                    if df_idx == p_vol_top3[0]: cell.fill, cell.font = FILLS["PINK1"], BLACK_FONT
                    elif len(p_vol_top3) > 1 and df_idx == p_vol_top3[1]: cell.fill, cell.font = FILLS["PINK2"], WHITE_FONT
                    elif len(p_vol_top3) > 2 and df_idx == p_vol_top3[2]: cell.fill, cell.font = FILLS["PINK3"], WHITE_FONT

        col_widths = [9,10,12,14,8,9,6,8,14,12,10,9]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"{index_name}_OC_{time.strftime('%Y%m%d_%H%M')}.xlsx"
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": TELEGRAM_CHAT_ID, "caption": f"📊 {index_name} Option Chain | {time.strftime('%d-%b %H:%M')}"},
            files={"document": (fname, buf.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=15
        )
    except Exception as e:
        pass


def send_telegram_strikewise(index_name, ltp, atm, pcr, df, step):
    try:
        msg_lines = [f"📊 *{index_name}* | LTP: `{ltp:,.0f}`\n"]
        df = df.copy()
        df["TOTAL_VOL"] = df["_cv"] + df["_pv"]
        df_top3 = df.nlargest(3, "TOTAL_VOL")
        df_sorted = df_top3.sort_values("STRIKE", ascending=False)

        def short_lakh(val): return f"{val / 1e5:.1f}L"

        for _, r in df_sorted.iterrows():
            strike = int(r["STRIKE"])
            c_vol_raw, p_vol_raw = r["_cv"], r["_pv"]

            icon = "🟢" if c_vol_raw > p_vol_raw else ("🔴" if p_vol_raw > c_vol_raw else "⚪")
            strike_txt = f"{icon} {strike} ATM" if strike == atm else f"{icon} {strike}"

            c_vol, p_vol = short_lakh(c_vol_raw), short_lakh(p_vol_raw)
            c_ltp, p_ltp = f"{float(r['C LTP']):.0f}", f"{float(r['P LTP']):.0f}"

            line = f"`{c_vol}/{c_ltp:<3}`  {strike_txt:^14}  `{p_ltp:>3}/{p_vol}`"
            msg_lines.append(line)

        msg_lines.append(f"\nPCR: `{pcr:.2f}`")
        requests.post(f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
                      json={"chat_id": TELEGRAM_CHAT_ID, "text": "\n".join(msg_lines), "parse_mode": "Markdown"}, timeout=10)
    except Exception as e:
        print(f"Error in send_telegram_strikewise: {e}")

# ──────────────────────────────────────────────
# SESSION INTERVAL BACKGROUND REFRESH
# ──────────────────────────────────────────────
def fmt_lakh(val): return f"{val/1e5:.1f}"

if "index_choice" not in st.session_state:
    st.session_state.index_choice = "NIFTY"

refresh_interval = 420
if "last_refresh" not in st.session_state: 
    st.session_state.last_refresh = time.time()

elapsed = time.time() - st.session_state.last_refresh
if elapsed >= refresh_interval:
    st.session_state.last_refresh = time.time()
    st.rerun()

col_btn1, col_btn2, col_spacer = st.columns([1, 1, 5])
with col_btn1:
    if st.button("NIFTY"):
        st.session_state.index_choice = "NIFTY"
        st.rerun()
with col_btn2:
    if st.button("SENSEX"):
        st.session_state.index_choice = "SENSEX"
        st.rerun()

cfg = UNDERLYING_MAP[st.session_state.index_choice]

# ──────────────────────────────────────────────
# CORE ZERODHA API DATA PROCESSING PIPELINE
# ──────────────────────────────────────────────
if kite:
    try:
        # 1. Fetch Underlying LTP to discover the active ATM boundary zone
        indices_quote = kite.quote(cfg["UnderlyingSymbol"])
        ltp = float(indices_quote[cfg["UnderlyingSymbol"]]["last_price"])
        atm = round(ltp / cfg["step"]) * cfg["step"]

        # 2. Fetch Daily Instruments Master Dump for NFO Contract matching
        @st.cache_data(ttl=3600)
        def fetch_nfo_instruments():
            instruments_df = pd.DataFrame(kite.instruments("NFO"))
            instruments_df['expiry'] = pd.to_datetime(instruments_df['expiry']).dt.date
            return instruments_df

        inst_df = fetch_nfo_instruments()
        filtered_ins = inst_df[inst_df["name"] == cfg["ExchangeSymbol"]]

        if not filtered_ins.empty:
            # Isolate the nearest chronological expiration date contract
            near_expiry = sorted(filtered_ins["expiry"].unique())[0]
            expiry_str = near_expiry.strftime("%Y-%m-%d")
            
            # Slice window around current target ATM domain (+/- 10 steps)
            lower_bound_strike = atm - (cfg["step"] * 10)
            upper_bound_strike = atm + (cfg["step"] * 10)
            
            chain_instruments = filtered_ins[
                (filtered_ins["expiry"] == near_expiry) & 
                (filtered_ins["strike"] >= lower_bound_strike) & 
                (filtered_ins["strike"] <= upper_bound_strike)
            ]

            # 3. Pull Batch market depth metrics via Quote block payload chunks
            trading_symbols = chain_instruments["tradingsymbol"].apply(lambda x: f"NFO:{x}").tolist()
            chunks = [trading_symbols[i:i + 50] for i in range(0, len(trading_symbols), 50)]
            quotes = {}
            for chunk in chunks:
                quotes.update(kite.quote(chunk))

            # Structuring matrix logic processing maps
            strikes_data = {}
            for _, inst in chain_instruments.iterrows():
                stk = float(inst["strike"])
                inst_type = inst["instrument_type"]
                sym = f"NFO:{inst['tradingsymbol']}"
                
                if stk not in strikes_data:
                    strikes_data[stk] = {"ce": {}, "pe": {}}
                
                if sym in quotes:
                    q = quotes[sym]
                    # Math formulation for tracing absolute day open interest variations safely
                    oi_val = q.get("oi", 0)
                    chg_pct = q.get("change", 0)
                    oi_delta = oi_val - (oi_val / (1 + (chg_pct/100) if chg_pct != -100 else 1))
                    
                    strikes_data[stk][inst_type.lower()] = {
                        "oi": oi_val,
                        "volume": q.get("volume", 0),
                        "last_price": q.get("last_price", 0),
                        "oi_change": int(oi_delta)
                    }

            # 4. Generate Structured DataFrame rows matching requested schema
            rows = []
            for strike_f, legs in strikes_data.items():
                ce, pe = legs.get("ce", {}), legs.get("pe", {})
                
                c_oi = int(ce.get("oi", 0))
                p_oi = int(pe.get("oi", 0))
                c_vol = int(ce.get("volume", 0))
                p_vol = int(pe.get("volume", 0))
                c_delta = int(ce.get("oi_change", 0))
                p_delta = int(pe.get("oi_change", 0))
                
                rows.append({
                    "C OI CH%": "0.0%",  # Kite quotes do not give static daily baseline parameters natively
                    "C VOL (L)": f"{c_vol/1e5:.2f}",
                    "CALL OI (L)": f"{c_oi/1e5:.2f}",
                    "C Δ OI": f"{c_delta:,} {'▲' if c_delta >= 0 else '▼'}",
                    "C LTP": f"{float(ce.get('last_price', 0)):.1f}",
                    "STRIKE": strike_f,
                    "IV": "0.0", 
                    "P LTP": f"{float(pe.get('last_price', 0)):.1f}",
                    "P Δ OI": f"{p_delta:,} {'▲' if p_delta >= 0 else '▼'}",
                    "PUT OI (L)": f"{p_oi/1e5:.2f}",
                    "P VOL (L)": f"{p_vol/1e5:.2f}",
                    "P OI CH%": "0.0%",
                    "_cv": c_vol, "_pv": p_vol,
                    "_cd": c_delta, "_pd": p_delta,
                    "_coi": c_oi, "_poi": p_oi
                })

            df = pd.DataFrame(rows).sort_values("STRIKE").reset_index(drop=True)
            
            # Compute analytical boundaries
            total_c_oi = df["_coi"].sum()
            total_p_oi = df["_poi"].sum()
            pcr = total_p_oi / total_c_oi if total_c_oi else 0

            c_vol_top3   = df['_cv'].nlargest(3).index.tolist()
            c_oi_top3    = df['_cd'].nlargest(3).index.tolist()
            p_vol_top3   = df['_pv'].nlargest(3).index.tolist()
            p_oi_top3    = df['_pd'].nlargest(3).index.tolist()
            min_c_oi_idx = df['_cd'].idxmin()
            min_p_oi_idx = df['_pd'].idxmin()
            c_neg_oi_top3 = df[df['_cd'] < 0]['_cd'].nsmallest(3).index.tolist()
            p_neg_oi_top3 = df[df['_pd'] < 0]['_pd'].nsmallest(3).index.tolist()

            # Execute alerts
            send_telegram_alert(st.session_state.index_choice, ltp, atm, expiry_str, pcr, df)
            send_excel_to_telegram(st.session_state.index_choice, ltp, atm, expiry_str, pcr, df, c_vol_top3, c_oi_top3, p_vol_top3, p_oi_top3, min_c_oi_idx, min_p_oi_idx, c_neg_oi_top3, p_neg_oi_top3)
            send_telegram_strikewise(st.session_state.index_choice, ltp, atm, pcr, df, cfg["step"])
            send_telegram_strikewise_image(st.session_state.index_choice, ltp, atm, pcr, df, cfg["step"])

            # Map dual analysis logic tracking transformations
            df_for_telegram = df.rename(columns={"STRIKE": "STRIKE", "CE Volume": "_cv", "PE Volume": "_pv", "CE Δ OI": "_cd", "PE Δ OI": "_pd"})
            # send_telegram_combined_analysis(st.session_state.index_choice, ltp, atm, pcr, df_for_telegram, cfg["step"])

            chart_image = render_strikewise_image_streamlit(st.session_state.index_choice, ltp, atm, pcr, df, cfg["step"])
            if chart_image is not None:
                st.image(chart_image, caption=f"{st.session_state.index_choice} Volume Chart", use_container_width=True)

            # ──────────────────────────────────────────────
            # METRICS BANNER DISPLAYS
            # ──────────────────────────────────────────────
            st.markdown(f"""
            <div style="background-color: #daeaf8; padding: 10px 0px; border-bottom: 1px solid #7ab3e0;">
                <h1 style="color: #0d1b2a; font-size: 2.2rem; font-weight: 800; margin-bottom: 20px; letter-spacing: -1px;">
                    NSE {st.session_state.index_choice} | ATM {int(atm)} | LTP {ltp:,.0f} | {time.strftime('%H:%M')}
                </h1>
                <div style="display: grid; grid-template-columns: repeat(6, 1fr); gap: 15px;">
                    <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">LTP</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{ltp:,.0f}</div></div>
                    <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">ATM</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{int(atm)}</div></div>
                    <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">PCR</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{pcr:.2f}</div></div>
                    <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">CE OI</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{fmt_lakh(total_c_oi)}L</div></div>
                    <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">PE OI</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{fmt_lakh(total_p_oi)}L</div></div>
                    <div><div style="color: #2c5f8a; font-size: 0.75rem; font-weight: 600;">CE OI Chg</div><div style="font-size: 1.6rem; font-weight: 700; color: #0d1b2a;">{fmt_lakh(df['_cd'].sum())}L</div></div>
                </div>
                <div style="color: #3a6ea5; font-size: 0.7rem; margin-top: 20px;">
                    Expiry: {expiry_str} | Update in: {int(refresh_interval - elapsed)}s
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("<div style='margin-bottom: 20px;'></div>", unsafe_allow_html=True)
            st.markdown("<div class='section-headers'><div class='sh'>CALLS</div><div class='sh'>STRIKE</div><div class='sh'>PUTS</div></div>", unsafe_allow_html=True)

            csv_data = df[["C OI CH%","C VOL (L)","CALL OI (L)","C Δ OI","C LTP","STRIKE","IV","P LTP","P Δ OI","PUT OI (L)","P VOL (L)","P OI CH%"]].to_csv(index=False)
            st.download_button(label="⬇️ Download CSV", data=csv_data, file_name=f"{st.session_state.index_choice}_OC_{time.strftime('%Y%m%d_%H%M')}.csv", mime="text/csv")

            # ──────────────────────────────────────────────
            # MATRIX TERMINAL CELL HIGHLIGHT LOGIC
            # ──────────────────────────────────────────────
            def style_terminal(data):
                styles = pd.DataFrame('', index=data.index, columns=data.columns)
                styles.update(pd.DataFrame('background-color: #f0f6ff; color: #0d1b2a;', index=data.index, columns=data.columns))
                styles['STRIKE'] = 'background-color: #c8dff5; color: #0d1b2a; font-weight: 700;'
                styles['IV']     = 'background-color: #daeaf8; color: #3a6ea5;'

                CYAN1, CYAN2, CYAN3 = '#1976d2', '#64b5f6', '#bbdefb'
                PINK1, PINK2, PINK3 = '#c62828', '#ef5350', '#ffcdd2'
                YELLOW1, YELLOW2, YELLOW3 = '#ffe082', '#ffd54f', '#fff9c4'

                for rank, (idx, bg) in enumerate(zip(c_vol_top3, [CYAN1, CYAN2, CYAN3])):
                    fg = '#000000' if rank == 0 else '#ffffff'
                    styles.loc[idx, 'C VOL (L)'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'
                for rank, (idx, bg) in enumerate(zip(c_oi_top3, [CYAN1, CYAN2, CYAN3])):
                    fg = '#000000' if rank == 0 else '#ffffff'
                    styles.loc[idx, 'C Δ OI'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'
                for rank, (idx, bg) in enumerate(zip(c_neg_oi_top3, [YELLOW1, YELLOW2, YELLOW3])):
                    styles.loc[idx, 'C Δ OI'] = f'background-color: {bg}; color: #000000; font-weight: 700;'
                for rank, (idx, bg) in enumerate(zip(p_vol_top3, [PINK1, PINK2, PINK3])):
                    fg = '#000000' if rank == 0 else '#ffffff'
                    styles.loc[idx, 'P VOL (L)'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'
                for rank, (idx, bg) in enumerate(zip(p_oi_top3, [PINK1, PINK2, PINK3])):
                    fg = '#000000' if rank == 0 else '#ffffff'
                    styles.loc[idx, 'P Δ OI'] = f'background-color: {bg}; color: {fg}; font-weight: 700;'
                for rank, (idx, bg) in enumerate(zip(p_neg_oi_top3, [YELLOW1, YELLOW2, YELLOW3])):
                    styles.loc[idx, 'P Δ OI'] = f'background-color: {bg}; color: #000000; font-weight: 700;'

                atm_idx = data[data['STRIKE'] == atm].index
                if not atm_idx.empty:
                    styles.loc[atm_idx[0], 'STRIKE'] = 'background-color: #ffffff; color: #000000; font-weight: 900;'
                return styles

            display_cols = ["C OI CH%", "C VOL (L)", "CALL OI (L)", "C Δ OI", "C LTP", "STRIKE", "IV", "P LTP", "P Δ OI", "PUT OI (L)", "P VOL (L)", "P OI CH%"]
            raw_cols = ["_cv", "_pv", "_cd", "_pd", "_coi", "_poi"]

            st.dataframe(
                df[display_cols + raw_cols].style.apply(style_terminal, axis=None)
                .format(precision=0).hide(axis="columns", subset=raw_cols),
                use_container_width=True, height=780
            )
        else:
            st.warning("No active contract instruments found in the NFO database dump.")
    except Exception as e:
        st.error(f"Error executing active workspace tracking pipeline: {e}")
else:
    st.info("💡 Please provide your active Zerodha API Key and Access Token in the sidebar console to initiate live processing operations.")

# Component forced script refresh 
st.components.v1.html(f"<script>setTimeout(function(){{ window.parent.location.reload(); }}, {refresh_interval * 1000});</script>", height=0)